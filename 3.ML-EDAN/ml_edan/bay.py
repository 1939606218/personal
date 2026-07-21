import torch
import os
from pre_data import loadData, pre_data_loader, normalize
from Global import *
from sst import SSTViT
# from main_train_test_CDFormer import train_test
# from make_graph_CDFormer import make_bay_graph
from main_train_test_ml_edan import train_test
from make_graph_ml_edan import make_bay_graph
# from try_bay import load_bay, pre_Bay_data_loader
from scipy import io
from einops import rearrange
from pre_data import set_train_sample
import torch.nn.functional as F
import matplotlib.pyplot as plt
import math
import openpyxl
import pandas as pd
from pre_Bay import patch_data, get_location, pre_bay_input_dbs3tan, pre_data_bay_loader, apply_pca

os.environ['CUDA_VISIBLE_DEVICES'] = '0, 1, 2, 3'
device = torch.device('cuda:0')


def main():
    wb = openpyxl.Workbook()
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    img_1, img_2, label = loadData(dataset)
    # x_1 = normalize(X=img_1, type=2)  # 选择标准化或者归一化
    # x_2 = normalize(X=img_2, type=2)
    # img_1 = apply_pca(img_1, num_components=pca_band)  # pca 降维 光谱波段冗余
    # img_2 = apply_pca(img_2, num_components=pca_band)
    location, labels = get_location(label) # 获取不为零位置
    a = patch_data(img_1, windowSize, location)
    b = patch_data(img_2, windowSize, location)     # 划分两张图为patch 只有非零位置
    train_inter, test_inter, total_inter = pre_data_bay_loader(a, b, labels)
    # train_inter, test_inter, total_inter = pre_bay_input_dbs3tan(a, b, labels)
    h, w, ch = img_1.shape  # H，W, C
    print(img_1.shape)
    # make_graph(h=h, w=w, device=device, total_inter=total_inter)

    for i in range(10):
        oa, kappa, f1 = train_test(device=device, train_inter=train_inter, test_inter=test_inter)
        make_bay_graph(h=h, w=w, device=device, total_inter=total_inter, location=location)

        name = ['OA', 'kappa*100', 'f1*100']
        number = [oa, kappa, f1]
        for j in range(len(number)):
            number[j] = round(number[j], 3)

        if i == 0:
            df = pd.DataFrame({'名称': name, '数据': number})
            ws = wb.create_sheet(title='Result')
            for index, row in df.iterrows():
                ws.cell(row=index + 1, column=1, value=row['名称'])
                ws.cell(row=index + 1, column=2, value=row['数据'])
        else:
            df = pd.DataFrame({'数据' + str(i): number})
            ws = wb['Result']
            for index, row in df.iterrows():
                ws.cell(row=index + 1, column=2 + i, value=row['数据' + str(i)])

    wb.save('Result' + '_' + str(dataset) + '_' + str(windowSize) + '_' + str(epochs) + '_' + str(pos) + '.xlsx')


if __name__ == '__main__':
    main()

