import json
import pandas as pd
import os
from pathlib import Path

def export_ablation_results_to_xlsx():
    """
    将消融实验结果导出到xlsx文件中
    格式：数据集 | 模块组合 | 各种指标
    从上到下依次是group 1-7，group7应该是精度最好的
    """
    
    # 定义数据集和对应的文件
    datasets = {
        'bayArea': 'ablation_results_bayArea_validated.json',
        'farmland': 'ablation_results_farmland_validated.json', 
        'santaBarbara': 'ablation_results_santaBarbara_validated.json'
    }
    
    # 定义组件映射
    component_mapping = {
        'SSDT': 'SSDT',
        'DRAF': 'DRAF', 
        'DCL': 'DCL',
        'HGLS': 'HGLS'
    }
    
    # 存储所有结果
    all_results = []
    
    # 处理每个数据集
    for dataset_name, filename in datasets.items():
        filepath = Path(__file__).parent / filename
        
        if not filepath.exists():
            print(f"警告: 文件 {filename} 不存在")
            continue
            
        # 读取JSON文件
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 按照group顺序处理 (group_1 到 group_7)
        for group_num in range(1, 8):
            group_key = f"group_{group_num}"
            
            if group_key not in data:
                continue
                
            group_data = data[group_key]
            
            # 构建模块组合字符串
            components = group_data.get('components', [])
            component_str = ' + '.join(components)
            
            # 获取指标
            metrics = group_data.get('metrics', {})
            
            # 创建结果行
            result_row = {
                'Dataset': dataset_name,
                'Group': f'Group {group_num}',
                'Components': component_str,
                'SSDT': '✓' if 'SSDT' in components else '✗',
                'DRAF': '✓' if 'DRAF' in components else '✗', 
                'DCL': '✓' if 'DCL' in components else '✗',
                'HGLS': '✓' if 'HGLS' in components else '✗',
                'OA': round(metrics.get('oa', 0), 4),
                'Kappa': round(metrics.get('kappa', 0), 4),
                'F1': round(metrics.get('f1', 0), 4),
                'Precision': round(metrics.get('pr', 0), 4),
                'Recall': round(metrics.get('re', 0), 4)
            }
            
            # 如果是baseline (group_7)，标记
            if group_data.get('is_baseline', False):
                result_row['Note'] = 'Baseline (Best)'
            else:
                result_row['Note'] = ''
                
            all_results.append(result_row)
    
    # 创建DataFrame
    df = pd.DataFrame(all_results)
    
    # 重新排序，确保每个数据集的group_7在最后
    df_sorted = pd.DataFrame()
    for dataset in ['bayArea', 'farmland', 'santaBarbara']:
        dataset_df = df[df['Dataset'] == dataset].copy()
        if not dataset_df.empty:
            df_sorted = pd.concat([df_sorted, dataset_df], ignore_index=True)
    
    # 保存到xlsx文件
    output_file = Path(__file__).parent / 'ablation_results_summary.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 保存主要结果
        df_sorted.to_excel(writer, sheet_name='Ablation Results', index=False)
        
        # 为每个数据集创建单独的sheet
        for dataset in ['bayArea', 'farmland', 'santaBarbara']:
            dataset_df = df_sorted[df_sorted['Dataset'] == dataset].copy()
            if not dataset_df.empty:
                # 移除Dataset列，因为sheet名已经表明了数据集
                dataset_df_clean = dataset_df.drop('Dataset', axis=1)
                sheet_name = dataset.replace('_', ' ').title()
                dataset_df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 创建性能对比汇总表
        summary_data = []
        for dataset in ['bayArea', 'farmland', 'santaBarbara']:
            dataset_df = df_sorted[df_sorted['Dataset'] == dataset]
            if not dataset_df.empty:
                # 找到baseline (group_7)
                baseline = dataset_df[dataset_df['Note'].str.contains('Baseline', na=False)]
                if not baseline.empty:
                    baseline_row = baseline.iloc[0]
                    summary_data.append({
                        'Dataset': dataset,
                        'Best Group': baseline_row['Group'],
                        'Components': baseline_row['Components'],
                        'OA': baseline_row['OA'],
                        'Kappa': baseline_row['Kappa'],
                        'F1': baseline_row['F1'],
                        'Precision': baseline_row['Precision'],
                        'Recall': baseline_row['Recall']
                    })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Best Results Summary', index=False)
    
    print(f"消融实验结果已成功导出到: {output_file}")
    print(f"共处理了 {len(df_sorted)} 组实验结果")
    print(f"涵盖数据集: {', '.join(datasets.keys())}")
    
    # 显示每个数据集的最佳结果
    print("\n各数据集最佳结果 (Group 7 - Baseline):")
    print("-" * 80)
    for dataset in ['bayArea', 'farmland', 'santaBarbara']:
        dataset_group7 = df_sorted[(df_sorted['Dataset'] == dataset) & 
                                  (df_sorted['Group'] == 'Group 7')]
        if not dataset_group7.empty:
            row = dataset_group7.iloc[0]
            print(f"{dataset}: OA={row['OA']:.4f}, Kappa={row['Kappa']:.4f}, "
                  f"F1={row['F1']:.4f}, Precision={row['Precision']:.4f}, Recall={row['Recall']:.4f}")

def format_excel_sheets():
    """
    格式化Excel表格，使其更美观
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    output_file = Path(__file__).parent / 'ablation_results_summary.xlsx'
    
    if not output_file.exists():
        print("Excel文件不存在，请先运行export_ablation_results_to_xlsx()")
        return
    
    wb = load_workbook(output_file)
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    baseline_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 格式化每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 设置列宽
        column_widths = {
            'A': 15, 'B': 12, 'C': 25, 'D': 8, 'E': 8, 'F': 8, 'G': 8,
            'H': 10, 'I': 10, 'J': 10, 'K': 12, 'L': 10, 'M': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 格式化标题行
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 格式化数据行
        for row_num in range(2, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = center_alignment
                
                # 如果是baseline行，设置背景色
                note_cell = ws.cell(row=row_num, column=ws.max_column)
                if note_cell.value and 'Baseline' in str(note_cell.value):
                    cell.fill = baseline_fill
    
    wb.save(output_file)
    print(f"Excel表格格式化完成: {output_file}")

if __name__ == "__main__":
    try:
        # 导出数据
        export_ablation_results_to_xlsx()
        
        # 格式化表格
        print("\n正在格式化Excel表格...")
        format_excel_sheets()
        
        print("\n✅ 消融实验结果导出完成！")
        print("📁 输出文件: ablation_results_summary.xlsx")
        print("📊 包含以下工作表:")
        print("   - Ablation Results: 完整的消融实验结果")
        print("   - Bay Area: bayArea数据集结果")
        print("   - Farmland: farmland数据集结果") 
        print("   - Santa Barbara: santaBarbara数据集结果")
        print("   - Best Results Summary: 各数据集最佳结果汇总")
        
    except Exception as e:
        print(f"❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()
